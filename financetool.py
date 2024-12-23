import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Optional, Union
import sys
import math
from decimal import Decimal, ROUND_HALF_UP
from types import SimpleNamespace
import json
import numpy as np
import os

class FinancialInputs:
    def __init__(self):
        self.initial_age: int = 0
        self.retirement_age: int = 0
        self.life_expectancy: int = 80
        self.initial_monthly_expense: float = 0
        self.monthly_emi: float = 0
        self.monthly_in_hand: float = 0  
        self.initial_pf_corpus: float = 0
        self.monthly_pf_contribution: float = 0
        self.initial_current_corpus: float = 0
        self.goals: List[Dict[str, Union[float, int]]] = []

    def validate(self) -> bool:
        if self.initial_age >= self.retirement_age:
            raise ValueError("Current age must be less than retirement age")
        if self.initial_age < 18:
            raise ValueError("Current age must be at least 18")
        if self.retirement_age >= self.life_expectancy:
            raise ValueError("Retirement age must be less than life expectancy")
        if self.initial_monthly_expense <= 0:
            raise ValueError("Monthly expenses must be greater than 0")
        if self.monthly_emi < 0:
            raise ValueError("Monthly EMI cannot be negative")
        if self.monthly_in_hand <= 0:
            raise ValueError("Monthly in-hand salary must be greater than 0")
        if self.initial_pf_corpus < 0:
            raise ValueError("PF corpus cannot be negative")
        if self.monthly_pf_contribution < 0:
            raise ValueError("Monthly PF contribution cannot be negative")
        if self.monthly_pf_contribution > self.monthly_in_hand:
            raise ValueError("Monthly PF contribution cannot be greater than monthly in-hand salary")
        if self.initial_current_corpus < 0:
            raise ValueError("Current corpus cannot be negative")
        if len(self.goals) > 10:
            raise ValueError("Maximum 10 goals are allowed")
        # Calculate total monthly outflow
        total_monthly_outflow = (
            self.initial_monthly_expense + 
            self.monthly_emi
        )
        if total_monthly_outflow > self.monthly_in_hand:
            raise ValueError("Total monthly outflow (expenses + EMI) cannot exceed monthly income")        
        return True

class FinancialConfig:
    def __init__(self, inputs: FinancialInputs):
        self.initial_age: int = inputs.initial_age
        self.retirement_age: int = inputs.retirement_age
        self.life_expectancy: int = inputs.life_expectancy
        self.initial_monthly_expense: float = inputs.initial_monthly_expense
        self.initial_annual_expense: float = inputs.initial_monthly_expense * 12
        self.expense_inflation_rate: float = 0.06
        self.investment_step_up_rate: float = 0.05
        self.rate_of_return: float = 0.12
        self.post_retirement_rate_of_return: float = 0.06

        # Asset allocation returns
        self.returns: Dict[str, float] = {
            "Equity": 0.16,
            "Gold": 0.13,
            "Real Estate": 0.08,
            "Debt": 0.08,
        }

        # Time-based goal returns
        self.get_goal_return = lambda years: (
            0.08 if years <= 3 else
            0.10 if years <= 5 else
            0.14
        )

class FinancialPlanner:
    def __init__(self, config: FinancialConfig):
        self.config = config

    def round_currency(self, amount: float) -> int:
        return int(Decimal(str(amount)).quantize(Decimal('1'), rounding=ROUND_HALF_UP))

    def calculate_retirement_plan(self, initial_investment: float, initial_pf_corpus: float = 0, 
                                monthly_pf_contribution: float = 0, initial_current_corpus: float = 0) -> pd.DataFrame:
        """
        Calculate retirement plan based on initial monthly investment and corpus values
        
        Args:
            initial_investment: Monthly investment amount
            initial_pf_corpus: Initial PF corpus amount (default: 0)
            monthly_pf_contribution: Monthly PF contribution (default: 0)
            initial_current_corpus: Initial current corpus amount (default: 0)
        
        Returns:
            pd.DataFrame: Retirement plan calculations
        """
        ages = list(range(self.config.initial_age, self.config.life_expectancy + 1))
        data = []
        
        monthly_exp = self.config.initial_monthly_expense
        annual_exp = self.config.initial_annual_expense
        investment = initial_investment
        salary_corpus = 0
        pf_corpus = initial_pf_corpus
        current_corpus = initial_current_corpus
        
        # Annual interest rates
        PF_INTEREST_RATE = 0.08  # 8% for PF
        CURRENT_INTEREST_RATE = 0.06  # 6% for current corpus

        for age in ages:
            year_data = {
                'Age': age,
                'Monthly Expense (₹)': self.round_currency(monthly_exp),
                'Annual Expense (₹)': self.round_currency(annual_exp)
            }

            if age <= self.config.retirement_age:
                year_data.update({
                    'Monthly Investment (₹)': self.round_currency(investment),
                    'Annual Investment (₹)': self.round_currency(investment * 12)
                })

                # Update all corpus types
                salary_corpus += (investment * 12)
                salary_corpus *= (1 + self.config.rate_of_return)
                
                # Update PF corpus with monthly contributions and interest
                pf_corpus += (monthly_pf_contribution * 12) 
                pf_corpus *= (1 + PF_INTEREST_RATE)
                
                # Apply interest rates to current corpus
                current_corpus *= (1 + CURRENT_INTEREST_RATE)

                total_corpus = salary_corpus + pf_corpus + current_corpus

                year_data.update({
                    'Salary Corpus (₹)': self.round_currency(salary_corpus),
                    'PF Corpus (₹)': self.round_currency(pf_corpus),
                    'Current Corpus (₹)': self.round_currency(current_corpus),
                    'Total Corpus (₹)': self.round_currency(total_corpus),
                    'Retirement Year-End Corpus (₹)': None
                })
            else:
                year_data.update({
                    'Monthly Investment (₹)': None,
                    'Annual Investment (₹)': None,
                    'Salary Corpus (₹)': None,
                    'PF Corpus (₹)': None,
                    'Current Corpus (₹)': None
                })

                # Calculate total corpus
                total_corpus = salary_corpus + pf_corpus + current_corpus
                
                # Deduct annual expenses proportionally from all corpus types
                if total_corpus > 0:  # Avoid division by zero
                    expense_ratio = annual_exp / total_corpus
                    salary_corpus *= (1 - expense_ratio)
                    pf_corpus *= (1 - expense_ratio)
                    current_corpus *= (1 - expense_ratio)
                
                # Apply post-retirement growth to all corpus types
                salary_corpus *= (1 + self.config.post_retirement_rate_of_return)
                pf_corpus *= (1 + self.config.post_retirement_rate_of_return)
                current_corpus *= (1 + self.config.post_retirement_rate_of_return)
                
                total_corpus = salary_corpus + pf_corpus + current_corpus
                year_data['Retirement Year-End Corpus (₹)'] = self.round_currency(total_corpus)

            data.append(year_data)

            monthly_exp *= (1 + self.config.expense_inflation_rate)
            annual_exp = monthly_exp * 12
            if age < self.config.retirement_age:
                investment *= (1 + self.config.investment_step_up_rate)

        return pd.DataFrame(data)
    
    def determine_retirement_investment(self) -> tuple[pd.DataFrame, float]:
        """Determine required monthly investment for retirement"""
        initial_investment = 5000
        max_iterations = 100000
        min_final_corpus = 100000  

        for _ in range(max_iterations):
            df = self.calculate_retirement_plan(initial_investment)
            final_corpus = df['Retirement Year-End Corpus (₹)'].iloc[-1]

            if pd.notnull(final_corpus) and final_corpus > min_final_corpus:
                return df, initial_investment

            initial_investment += 1000

        raise ValueError("Could not find suitable investment amount within iterations limit")

    def calculate_investment_allocation(self, monthly_investment: float, target_corpus: float) -> pd.DataFrame:
        """
        Calculate investment allocation across different assets with dynamic adjustment 
        to ensure corpus progress reaches 100%
        
        Args:
            monthly_investment: Initial monthly investment amount
            target_corpus: Total target corpus to be achieved
        
        Returns:
            pd.DataFrame: Detailed investment allocation and progress
        """
        data = []
        prev_total_investment = 0
        current_monthly_investment = monthly_investment
        total_target_corpus = target_corpus

        def get_age_based_allocation(age: int) -> Dict[str, float]:
            if age <= 25:
                return {"Equity": 0.65, "Gold": 0.2, "Real Estate": 0.1, "Debt": 0.05}
            elif age <= 30:
                return {"Equity": 0.6, "Gold": 0.2, "Real Estate": 0.15, "Debt": 0.05}
            elif age <= 35:
                return {"Equity": 0.55, "Gold": 0.2, "Real Estate": 0.15, "Debt": 0.1}
            elif age <= 45:
                return {"Equity": 0.5, "Gold": 0.2, "Real Estate": 0.2, "Debt": 0.1}
            elif age <= 50:
                return {"Equity": 0.4, "Gold": 0.25, "Real Estate": 0.25, "Debt": 0.1}
            else:
                return {"Equity": 0.35, "Gold": 0.3, "Real Estate": 0.25, "Debt": 0.1}

        # Determine maximum investment adjustment iterations
        MAX_ITERATIONS = 50
        investment_multiplier = 1.0
        monthly_investments = []  # Track monthly investments

        for iteration in range(MAX_ITERATIONS):
            data = []
            monthly_investments = []  
            prev_total_investment = 0
            current_monthly_investment = monthly_investment * investment_multiplier
            accumulated_amount = 0

            for age in range(self.config.initial_age, self.config.retirement_age + 1):
                # Ensure the current monthly investment is a multiple of 500 (rounded up)
                current_monthly_investment = math.ceil(current_monthly_investment / 500) * 500
                
                # Accumulate yearly investment
                yearly_investment = self.round_currency(current_monthly_investment * 12)
                monthly_investments.append(self.round_currency(current_monthly_investment))
                
                # Determine the allocations based on age
                allocations = get_age_based_allocation(age)

                # Calculate the investment for each asset
                investments = {
                    asset: self.round_currency(yearly_investment * alloc)
                    for asset, alloc in allocations.items()
                }

                # Calculate the returns for each asset based on the previous total investment
                asset_returns = {
                    asset: self.round_currency(prev_total_investment * alloc * self.config.returns[asset])
                    for asset, alloc in allocations.items()
                }

                # Total return this year
                total_return = sum(asset_returns.values())
                total_return_percentage = (total_return / prev_total_investment * 100) if prev_total_investment > 0 else 0

                # Accumulated total corpus with investment and returns
                accumulated_amount = prev_total_investment + total_return + yearly_investment

                # Ensure the accumulated corpus does not exceed the target corpus
                accumulated_amount = min(accumulated_amount, total_target_corpus)

                # Calculate the corpus progress percentage
                corpus_progress_percentage = (accumulated_amount / total_target_corpus) * 100

                # Create a dictionary for the current year data
                year_data = {
                    "Age": age,
                    "Monthly Investment (₹)": self.round_currency(current_monthly_investment),
                    "Yearly Investment (₹)": yearly_investment,
                    **{f"{asset} Allocation (%)": round(alloc * 100, 1) for asset, alloc in allocations.items()},
                    **{f"{asset} Investment (₹)": investments[asset] for asset in allocations},
                    **{f"{asset} Return (₹)": asset_returns[asset] for asset in allocations},
                    "Total Return %": round(total_return_percentage, 2),
                    "Accumulated Investment with Compound Interest (₹)": self.round_currency(accumulated_amount),
                    "Target Corpus (₹)": self.round_currency(total_target_corpus),
                    "Corpus Progress %": round(corpus_progress_percentage, 2)
                }

                # Add the year data to the result list
                data.append(year_data)

                # Update the previous total investment for the next iteration
                prev_total_investment = accumulated_amount

                # Step up the monthly investment for the next year
                current_monthly_investment *= (1 + self.config.investment_step_up_rate)

            # Check if we've reached 100% corpus progress
            final_corpus_progress = data[-1]["Corpus Progress %"]
            
            # Fine-tune the investment multiplier
            if final_corpus_progress >= 99.9 and final_corpus_progress <= 100.1:
                break
            elif final_corpus_progress < 99.9:
                # If progress is too low, increase investment
                investment_multiplier *= 1.05
            else:
                # If progress is too high, decrease investment
                investment_multiplier *= 0.95

        # Return the DataFrame containing the investment allocation details
        return pd.DataFrame(data), monthly_investments

    def get_goal_allocation(self, age: int, years: int) -> Dict[str, float]:
        """Determine asset allocation based on age and investment horizon"""
        if years <= 3:
            return {'Equity': 0.0, 'Gold': 0.0, 'Debt': 1.0}
        elif years <= 5:
            if age <= 35:
                return {'Equity': 0.6, 'Gold': 0.0, 'Debt': 0.4}
            else:
                return {'Equity': 0.5, 'Gold': 0.0, 'Debt': 0.5}
        else:
            if age <= 35:
                return {'Equity': 0.7, 'Gold': 0.1, 'Debt': 0.2}
            elif age <= 50:
                return {'Equity': 0.5, 'Gold': 0.3, 'Debt': 0.2}
            else:
                return {'Equity': 0.2, 'Gold': 0.5, 'Debt': 0.3}


    def calculate_goal_investment(self, years: int, target_amount: float) -> pd.DataFrame:
        """
        Calculate investment required for financial goals with yearly step-up logic and proper allocation.
        Includes detailed asset allocation and progress tracking.
        """
        # Step 1: Calculate the target return rate for the given number of years
        target_return = self.config.get_goal_return(years)

        # Step 2: Calculate initial yearly investment for the target amount (without step-up applied yet)
        compound_factor = (1 + target_return) ** years
        initial_yearly_investment = target_amount / ((compound_factor - 1) / target_return)  
        initial_monthly_investment = initial_yearly_investment / 12

        # Initialize lists for storing data
        data = []
        accumulated_investment = 0
        prev_total_investment = 0

        # Track monthly and yearly investments with step-up
        current_yearly_investment = initial_yearly_investment
        current_monthly_investment = initial_monthly_investment

        # Step 3: Loop through each year dynamically
        for age in range(self.config.initial_age, self.config.initial_age + years):
            # Get asset allocation for the current year
            allocation = self.get_goal_allocation(age, years)

            # Apply step-up logic
            if age > self.config.initial_age:  
                current_yearly_investment *= (1 + self.config.investment_step_up_rate)
                current_monthly_investment = current_yearly_investment / 12
            
            # Round monthly investment to nearest ₹500
            current_monthly_investment = math.ceil(current_monthly_investment / 500) * 500
            current_yearly_investment = current_monthly_investment * 12

            # Calculate investments for each asset class
            investments = {
                asset: self.round_currency(current_yearly_investment * alloc)
                for asset, alloc in allocation.items()
            }

            # Calculate returns for each asset class based on accumulated investment
            if prev_total_investment > 0:
                returns = {
                    asset: self.round_currency(prev_total_investment * alloc * self.config.returns[asset])
                    for asset, alloc in allocation.items()
                }
                total_return = sum(returns.values())
                return_percentage = (total_return / prev_total_investment) * 100
            else:
                returns = {asset: 0 for asset in allocation}
                total_return = 0
                return_percentage = 0

            # Update accumulated investment
            prev_total_investment = accumulated_investment
            accumulated_investment += current_yearly_investment + total_return

            # Calculate corpus progress percentage
            corpus_progress = (accumulated_investment / target_amount) * 100

            # Store data for the current year
            year_data = {
                "Age": age,
                "Monthly Investment (₹)": self.round_currency(current_monthly_investment),
                "Yearly Investment (₹)": self.round_currency(current_yearly_investment),
                "Equity Allocation (%)": round(allocation.get('Equity', 0) * 100, 1),
                "Equity Investment (₹)": investments.get('Equity', 0),
                "Equity Return (₹)": returns.get('Equity', 0),
                "Gold Allocation (%)": round(allocation.get('Gold', 0) * 100, 1),
                "Gold Investment (₹)": investments.get('Gold', 0),
                "Gold Return (₹)": returns.get('Gold', 0),
                "Debt Allocation (%)": round(allocation.get('Debt', 0) * 100, 1),
                "Debt Investment (₹)": investments.get('Debt', 0),
                "Debt Return (₹)": returns.get('Debt', 0),
                "Total Return %": round(return_percentage, 2),
                "Accumulated Investment with Compound Interest (₹)": self.round_currency(accumulated_investment),
                "Target Corpus (₹)": self.round_currency(target_amount),
                "Corpus Progress %": round(corpus_progress, 2)
            }

            data.append(year_data)

        # Return DataFrame with investment data
        return pd.DataFrame(data)

def create_summary_sheet(
        inputs: FinancialInputs,
        age: int,
        investment_monthly_investments: List[float],
        goal_monthlies: List[float],
        monthly_in_hand: float,
        monthly_expenses: float,
        monthly_pf_contribution: float,
        monthly_emi: float,        
        final_retirement_corpus: float,
        years_to_retirement: int,
        investment_step_up_rate: float = 0.05,
        rate_of_return: float = 0.12,
        inflation_rate: float = 0.06
) -> pd.DataFrame:
    """Create summary sheet with monthly investment requirements, with and without step-up"""

    # Use the first month's investment as the retirement monthly investment
    retirement_monthly = investment_monthly_investments[0]

    # Calculate flat monthly investments for retirement and goals
    flat_retirement_monthly = calculate_flat_monthly_investment(
        retirement_monthly, 
        years_to_retirement, 
        rate_of_return, 
        investment_step_up_rate
    )
    
    flat_goal_monthlies = [
        calculate_flat_monthly_investment(
            monthly, 
            goal['years'], 
            rate_of_return, 
            investment_step_up_rate
        )
        for monthly, goal in zip(goal_monthlies, inputs.goals)
    ]

    # Calculate total monthly investment required (with step-up)
    total_monthly_required = retirement_monthly + sum(goal_monthlies)
    total_flat_monthly_required = flat_retirement_monthly + sum(flat_goal_monthlies)

    # Calculate monthly surplus/deficit including monthly expenses
    monthly_surplus = monthly_in_hand - total_monthly_required - monthly_expenses - monthly_emi
    flat_monthly_surplus = monthly_in_hand - total_flat_monthly_required - monthly_expenses - monthly_emi

    # Calculate present value of final corpus
    present_value = final_retirement_corpus / ((1 + inflation_rate) ** years_to_retirement)

    # Custom formatter for Indian Rupee
    def format_rupees(value: str) -> str:
        if not value or pd.isna(value):  # Handle empty or NaN values
            return ''
        
        # Convert the value to a number (float or int)
        try:
            numeric_value = float(value)
        except ValueError:
            return ''  # If the value is not a valid number, return an empty string
        
        # Convert to string and split integer and decimal parts
        str_value = f'{numeric_value:.0f}'
        
        # Format with Indian-style comma separators
        if len(str_value) <= 3:
            return f'₹{str_value}'
        
        # Separate the first part (before the first comma) and the rest
        first_part = str_value[:-3]  # All digits except the last 3
        last_part = str_value[-3:]   # The last 3 digits
        
        # Reverse the first part for easier comma insertion
        reversed_first_part = first_part[::-1]
        
        # Insert commas every 2 digits in the reversed first part
        formatted_first_part = ','.join([
            reversed_first_part[i:i+2] for i in range(0, len(reversed_first_part), 2)
        ])
        
        # Reverse the formatted first part back to correct order
        formatted_first_part = formatted_first_part[::-1]
        
        # Combine the formatted parts
        formatted_value = formatted_first_part + ',' + last_part
        
        return f'₹{formatted_value}'


    # Create summary data with new columns
    summary_data = {
        'Description': [
            'Current Age',
            'Monthly In-Hand Salary',
            'Current Monthly Expenses',
            'Monthly EMI',
            'Monthly PF Contribution',
            'Required Monthly Investment for Retirement (with Step-up)',
            'Required Monthly Investment for Retirement (Flat)',
            *[f'Required Monthly Investment for Goal {i + 1} (with Step-up)' for i in range(len(goal_monthlies))],
            *[f'Required Monthly Investment for Goal {i + 1} (Flat)' for i in range(len(goal_monthlies))],
            'Total Monthly Investment Required (with Step-up)',
            'Total Monthly Investment Required (Flat)',
            'Monthly Surplus/Deficit (with Step-up)',
            'Monthly Surplus/Deficit (Flat)',
            'Final Retirement Corpus (Future Value)',
            'Final Retirement Corpus (Present Value)'
        ],
        'Amount': [
            age,
            monthly_in_hand,
            monthly_expenses,
            monthly_emi,
            monthly_pf_contribution,
            retirement_monthly,
            flat_retirement_monthly,
            *goal_monthlies,
            *flat_goal_monthlies,
            total_monthly_required,
            total_flat_monthly_required,
            monthly_surplus,
            flat_monthly_surplus,
            final_retirement_corpus,
            present_value
        ]
    }

    # Convert numeric columns to DataFrame
    df = pd.DataFrame(summary_data)
    
    # Format numeric columns (except age) as Indian Rupee
    df['Formatted Amount'] = df.apply(
        lambda row: row['Amount'] if row['Description'] == 'Current Age' 
        else format_rupees(row['Amount']), 
        axis=1
    )
    
    # Drop the original 'Amount' column and rename 'Formatted Amount'
    df = df.drop('Amount', axis=1).rename(columns={'Formatted Amount': 'Amount'})

    return df

    # Calculate monthly investment without step-up
def calculate_flat_monthly_investment(
    step_up_monthly: float, 
    years: int, 
    rate_of_return: float, 
    step_up_rate: float
) -> float:
    """
    Calculate flat monthly investment to achieve the same corpus as step-up investment
    
    Args:
        step_up_monthly: Monthly investment with step-up
        years: Investment duration
        rate_of_return: Expected annual return rate
        step_up_rate: Annual investment step-up rate
    
    Returns:
        Flat monthly investment amount
    """
    # Calculate future value of step-up investment
    step_up_corpus = 0
    current_monthly = step_up_monthly
    
    for year in range(years):
        # Annual investment with compound interest
        year_investment = current_monthly * 12
        step_up_corpus = (step_up_corpus + year_investment) * (1 + rate_of_return)
        
        # Increase monthly investment for next year
        current_monthly *= (1 + step_up_rate)
    
    # Calculate flat monthly investment to achieve same corpus
    def calculate_flat_corpus(flat_monthly):
        flat_corpus = 0
        for year in range(years):
            # Annual investment with compound interest
            year_investment = flat_monthly * 12
            flat_corpus = (flat_corpus + year_investment) * (1 + rate_of_return)
        return flat_corpus
    
    # Binary search to find equivalent flat monthly investment
    left, right = step_up_monthly, step_up_monthly * 2
    while right - left > 1:
        mid = (left + right) / 2
        mid_corpus = calculate_flat_corpus(mid)
        
        if mid_corpus < step_up_corpus:
            left = mid
        else:
            right = mid
    
    return (left + right) / 2

    # Custom formatter for Indian Rupee (same as before)
    def format_rupees(value: str) -> str:
        if not value or pd.isna(value):  # Handle empty or NaN values
            return ''
        
        # Convert the value to a number (float or int)
        try:
            numeric_value = float(value)
        except ValueError:
            return ''  # If the value is not a valid number, return an empty string
        
        # Convert to string and split integer and decimal parts
        str_value = f'{numeric_value:.0f}'
        
        # Format with Indian-style comma separators
        if len(str_value) <= 3:
            return f'₹{str_value}'
        
        # Separate the first part (before the first comma) and the rest
        first_part = str_value[:-3]  # All digits except the last 3
        last_part = str_value[-3:]   # The last 3 digits
        
        # Reverse the first part for easier comma insertion
        reversed_first_part = first_part[::-1]
        
        # Insert commas every 2 digits in the reversed first part
        formatted_first_part = ','.join([
            reversed_first_part[i:i+2] for i in range(0, len(reversed_first_part), 2)
        ])
        
        # Reverse the formatted first part back to correct order
        formatted_first_part = formatted_first_part[::-1]
        
        # Combine the formatted parts
        formatted_value = formatted_first_part + ',' + last_part
        
        return f'₹{formatted_value}'

    # Create summary data with new columns
    summary_data = {
        'Description': [
            'Current Age',
            'Monthly In-Hand Salary',
            'Current Monthly Expenses',
            'Monthly PF Contribution',
            'Required Monthly Investment for Retirement (with Step-up)',
            'Required Monthly Investment for Retirement (Flat)',
            *[f'Required Monthly Investment for Goal {i + 1} (with Step-up)' for i in range(len(goal_monthlies))],
            *[f'Required Monthly Investment for Goal {i + 1} (Flat)' for i in range(len(goal_monthlies))],
            'Total Monthly Investment Required (with Step-up)',
            'Total Monthly Investment Required (Flat)',
            'Monthly Surplus/Deficit (with Step-up)',
            'Monthly Surplus/Deficit (Flat)',
            'Final Retirement Corpus (Future Value)',
            'Final Retirement Corpus (Present Value)'
        ],
        'Amount': [
            age,
            monthly_in_hand,
            monthly_expenses,
            monthly_pf_contribution,
            retirement_monthly,
            flat_retirement_monthly,
            *goal_monthlies,
            *flat_goal_monthlies,
            total_monthly_required,
            total_flat_monthly_required,
            monthly_surplus,
            flat_monthly_surplus,
            final_retirement_corpus,
            present_value
        ]
    }

    # Convert numeric columns to DataFrame
    df = pd.DataFrame(summary_data)
    
    # Format numeric columns (except age) as Indian Rupee
    df['Formatted Amount'] = df.apply(
        lambda row: row['Amount'] if row['Description'] == 'Current Age' 
        else format_rupees(row['Amount']), 
        axis=1
    )
    
    # Drop the original 'Amount' column and rename 'Formatted Amount'
    df = df.drop('Amount', axis=1).rename(columns={'Formatted Amount': 'Amount'})

    return df

def create_excel_report(
        retirement_df: pd.DataFrame,
        investment_df: pd.DataFrame,
        goal_dfs: List[tuple[str, pd.DataFrame]],
        summary_df: pd.DataFrame,
        output_path: str
) -> bool:
    """Create Excel report with monthly investment columns"""
    try:
        wb = Workbook()
                
        def format_worksheet(worksheet, df, show_monthly=True):
            """Apply consistent formatting to worksheet with specific column order"""
            from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
            
            # Define the exact column order
            standard_columns = [
                'Age',
                'Monthly Investment (₹)',
                'Yearly Investment (₹)',
                'Equity Allocation (%)',
                'Equity Monthly Investment (₹)',
                'Equity Investment (₹)',
                'Equity Return (₹)',
                'Gold Allocation (%)',
                'Gold Monthly Investment (₹)',
                'Gold Investment (₹)',
                'Gold Return (₹)',
                'Real Estate Allocation (%)',
                'Real Estate Monthly Investment (₹)',
                'Real Estate Investment (₹)',
                'Real Estate Return (₹)',
                'Debt Allocation (%)',
                'Debt Monthly Investment (₹)',
                'Debt Investment (₹)',
                'Debt Return (₹)',
                'Total Return %',
                'Accumulated Investment with Compound Interest (₹)',
                'Target Corpus (₹)',
                'Corpus Progress %'
            ]
            
            # Prepare DataFrame with correct columns
            if show_monthly:
                # Add missing monthly columns
                for asset in ['Equity', 'Gold', 'Real Estate', 'Debt']:
                    yearly_col = f'{asset} Investment (₹)'
                    monthly_col = f'{asset} Monthly Investment (₹)'
                    if yearly_col in df.columns and monthly_col not in df.columns:
                        df[monthly_col] = df[yearly_col] / 12
            
            # Reorder columns according to standard format
            available_columns = [col for col in standard_columns if col in df.columns]
            extra_columns = [col for col in df.columns if col not in standard_columns]
            df = df[available_columns + extra_columns]
            
            # Reorder columns according to standard format
            available_columns = [col for col in standard_columns if col in df.columns]
            extra_columns = [col for col in df.columns if col not in standard_columns]
            df = df[available_columns + extra_columns]
            
            # Define styles
            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            header_font = Font(bold=True, size=11)
            normal_font = Font(size=10)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Write headers
            for idx, col in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=idx, value=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Set column width
                max_length = max(len(str(df[col].astype(str))), len(str(col)))
                adjusted_width = min(max_length + 4, 50)
                worksheet.column_dimensions[cell.column_letter].width = adjusted_width

            # Write and format data
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.font = normal_font
                    
                    # Format numbers
                    if isinstance(value, (int, float)):
                        if 'Investment' in df.columns[col_idx-1] or 'Corpus' in df.columns[col_idx-1]:
                            cell.number_format = '₹#,##0'
                            cell.alignment = Alignment(horizontal='right')
                        elif 'Allocation' in df.columns[col_idx-1]:
                            cell.number_format = '0.00%'
                            cell.alignment = Alignment(horizontal='right')
                        elif 'Progress' in df.columns[col_idx-1] or 'Return %' in df.columns[col_idx-1]:
                            cell.number_format = '0.00%'
                            cell.alignment = Alignment(horizontal='right')
                        else:
                            cell.alignment = Alignment(horizontal='right')
                    else:
                        cell.alignment = Alignment(horizontal='left')

            # Freeze header row and add auto-filter
            worksheet.freeze_panes = 'A2'
            worksheet.auto_filter.ref = worksheet.dimensions
            
            return df
        
        # Summary sheet
        ws = wb.active
        ws.title = 'Summary'
        format_worksheet(ws, summary_df, show_monthly=False)
        
        # Retirement Plan sheet
        retirement_sheet = wb.create_sheet('Retirement Plan')
        # Add monthly columns for retirement investments
        retirement_df = format_worksheet(retirement_sheet, retirement_df)
        
        # Investment Allocation sheet
        investment_sheet = wb.create_sheet('Investment Allocation')
        # Add monthly columns for each investment type
        investment_df = format_worksheet(investment_sheet, investment_df)
        
        # Goal sheets with monthly investments
        for goal_name, goal_df in goal_dfs:
            goal_sheet = wb.create_sheet(goal_name)
            # Add monthly columns for each investment type
            goal_df = format_worksheet(goal_sheet, goal_df)

        # Save workbook
        wb.save(output_path)
        print(f"Excel report successfully created at: {output_path}")
        return True
        
    except Exception as e:
        print(f"Error creating Excel report: {str(e)}")
        import traceback
        traceback.print_exc()
        return False
    

def prepare_investment_df(df: pd.DataFrame) -> pd.DataFrame:
    """Prepare investment DataFrame with standardized columns"""
    # Add monthly investment columns for each asset class
    for asset in ['Equity', 'Gold', 'Real Estate', 'Debt']:
        yearly_col = f'{asset} Investment (₹)'
        monthly_col = f'{asset} Monthly Investment (₹)'
        if yearly_col in df.columns:
            df[monthly_col] = df[yearly_col] / 12
    
    # Define standard column order
    standard_columns = [
        'Age',
        'Monthly Investment (₹)',
        'Yearly Investment (₹)',
        'Equity Allocation (%)',
        'Equity Monthly Investment (₹)',
        'Equity Investment (₹)',
        'Equity Return (₹)',
        'Gold Allocation (%)',
        'Gold Monthly Investment (₹)',
        'Gold Investment (₹)',
        'Gold Return (₹)',
        'Real Estate Allocation (%)',
        'Real Estate Monthly Investment (₹)',
        'Real Estate Investment (₹)',
        'Real Estate Return (₹)',
        'Debt Allocation (%)',
        'Debt Monthly Investment (₹)',
        'Debt Investment (₹)',
        'Debt Return (₹)',
        'Total Return %',
        'Accumulated Investment with Compound Interest (₹)',
        'Target Corpus (₹)',
        'Corpus Progress %'
    ]
    
    # Reorder columns
    available_columns = [col for col in standard_columns if col in df.columns]
    extra_columns = [col for col in df.columns if col not in standard_columns]
    df = df[available_columns + extra_columns]
    
    return df

def prepare_goal_df(df: pd.DataFrame) -> pd.DataFrame:
    """Prepare goal DataFrame with monthly columns and arrange columns in the desired order"""
    # Add monthly investment columns
    investment_cols = [col for col in df.columns if 'Investment (₹)' in col and 'Monthly' not in col]
    for col in investment_cols:
        monthly_col = col.replace('Investment (₹)', 'Monthly Investment (₹)')
        df[monthly_col] = df[col] / 12
    
    # Define the desired column order
    desired_columns = [
        'Age', 
        'Monthly Investment (₹)', 
        'Yearly Investment (₹)', 
        'Equity Allocation (%)', 
        'Equity Monthly Investment (₹)', 
        'Equity Investment (₹)', 
        'Equity Return (₹)', 
        'Gold Allocation (%)', 
        'Gold Monthly Investment (₹)', 
        'Gold Investment (₹)', 
        'Gold Return (₹)', 
        'Debt Allocation (%)', 
        'Debt Monthly Investment (₹)', 
        'Debt Investment (₹)', 
        'Debt Return (₹)', 
        'Total Return %', 
        'Accumulated Investment with Compound Interest (₹)', 
        'Target Corpus (₹)', 
        'Corpus Progress %'
    ]
    
    # Reorder the columns based on the desired order, keeping extra columns at the end
    ordered_columns = [col for col in desired_columns if col in df.columns]
    remaining_columns = [col for col in df.columns if col not in desired_columns]
    df = df[ordered_columns + remaining_columns]
    
    return df

def get_user_inputs_from_dict(data: dict) -> FinancialInputs:
    """Accept and validate user inputs from a dictionary."""
    try:
        # Create an instance of FinancialInputs
        inputs = FinancialInputs()

        # Personal Information
        inputs.initial_age = data['initial_age']
        inputs.retirement_age = data['retirement_age']
        inputs.initial_monthly_expense = data['monthly_expenses']
        inputs.monthly_in_hand = data['monthly_in_hand']

        # Add EMI to existing inputs
        inputs.monthly_emi = data.get('monthly_emi', 0)

        # Existing Corpus Information
        inputs.initial_pf_corpus = data['initial_pf_corpus']
        inputs.monthly_pf_contribution = data['monthly_pf_contribution']
        inputs.initial_current_corpus = data['initial_current_corpus']

        # Financial Goals
        num_goals = len(data.get('goals', []))
        if not 1 <= num_goals <= 10:
            raise ValueError("Number of goals must be between 1 and 10")

        for goal in data['goals']:
            inputs.goals.append({
                'amount': goal['amount'],
                'years': goal['years']
            })

        # Validate inputs
        inputs.validate()

        return inputs

    except KeyError as e:
        raise ValueError(f"Missing required input: {e}")
    except ValueError as e:
        raise ValueError(f"Invalid input: {e}")
    except Exception as e:
        raise Exception(f"Unexpected error: {e}")

def get_user_inputs() -> Optional[FinancialInputs]:
    """Get and validate all user inputs"""
    try:
        inputs = FinancialInputs()
        print("\n=== Personal Information ===")
        inputs.initial_age = int(input("Enter your current age: "))
        inputs.retirement_age = int(input("Enter your desired retirement age: "))
        inputs.initial_monthly_expense = float(input("Enter your current monthly expenses (in ₹): "))
        inputs.monthly_in_hand = float(input("Enter your monthly in-hand salary (in ₹): "))
        inputs.monthly_emi = float(input("Enter your Monthly EMI (in ₹) [Enter 0 if none]: "))

        print("\n=== Existing Corpus Information ===")
        inputs.initial_pf_corpus = float(input("Enter your current PF corpus amount (in ₹) [Enter 0 if none]: "))
        inputs.monthly_pf_contribution = float(input("Enter your monthly PF contribution (in ₹): "))
        inputs.initial_current_corpus = float(input("Enter your current investment corpus amount (in ₹) [Enter 0 if none]: "))

        print("\n=== Financial Goals ===")
        num_goals = int(input("Enter the number of financial goals you have (1-10): "))
        if not 1 <= num_goals <= 10:
            raise ValueError("Number of goals must be between 1 and 10")

        for i in range(num_goals):
            print(f"\nGoal {i + 1} Details:")
            amount = float(input(f"Enter target amount for Goal {i + 1} (in ₹): "))
            years = int(input(f"Enter number of years to achieve Goal {i + 1}: "))
            inputs.goals.append({'amount': amount, 'years': years})

        inputs.validate()
        return inputs

    except ValueError as e:
        print(f"Invalid input: {str(e)}")
        return None
    except Exception as e:
        print(f"Error: {str(e)}")
        return None
    
def convert_to_serializable(obj):
    """Convert non-serializable types to serializable equivalents."""
    if isinstance(obj, (np.int64, np.int32)):
        return int(obj)
    if isinstance(obj, (np.float64, np.float32)):
        return float(obj)
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    return obj

def main(data=None):
    try:
        print("\n=== Financial Planning Tool ===")
        
        # Create output directory if it doesn't exist
        output_dir = "static/generated_files"
        os.makedirs(output_dir, exist_ok=True)
        
        # Use data if provided; otherwise, get inputs interactively
        if data:
            inputs = get_user_inputs_from_dict(data)
        else:
            inputs = get_user_inputs()
        
        if inputs is None:
            return {"success": False, "message": "No valid inputs provided."}

        config = FinancialConfig(inputs)
        planner = FinancialPlanner(config)

        print("\nCalculating retirement plan...")
        retirement_df, required_monthly_investment = planner.determine_retirement_investment()

        # Get final corpus from retirement plan
        final_retirement_corpus = retirement_df['Total Corpus (₹)'].iloc[retirement_df['Total Corpus (₹)'].last_valid_index()]
        years_to_retirement = inputs.retirement_age - inputs.initial_age

        retirement_df = planner.calculate_retirement_plan(
            initial_investment=required_monthly_investment,
            initial_pf_corpus=inputs.initial_pf_corpus,
            monthly_pf_contribution=inputs.monthly_pf_contribution,
            initial_current_corpus=inputs.initial_current_corpus
        )

        print("Calculating investment allocation...")
        investment_df, monthly_investments = planner.calculate_investment_allocation(
            required_monthly_investment, 
            final_retirement_corpus
        )

        print("Calculating goal-based investments...")
        goal_dfs = []
        goal_monthly_investments = []
        for i, goal in enumerate(inputs.goals):
            goal_df = planner.calculate_goal_investment(goal['years'], goal['amount'])
            goal_monthly_investment = goal_df['Monthly Investment (₹)'].iloc[0]
            goal_dfs.append((f"Goal {i + 1} ({goal['years']} years)", goal_df))
            goal_monthly_investments.append(goal_monthly_investment)

        # Create summary with EMI included
        summary_df = create_summary_sheet(
            inputs=inputs,
            age=inputs.initial_age,
            investment_monthly_investments=monthly_investments,
            goal_monthlies=goal_monthly_investments,
            monthly_in_hand=inputs.monthly_in_hand,
            monthly_expenses=inputs.initial_monthly_expense,
            monthly_pf_contribution=inputs.monthly_pf_contribution,
            monthly_emi=inputs.monthly_emi,  # Include EMI
            final_retirement_corpus=final_retirement_corpus,
            years_to_retirement=years_to_retirement,
            investment_step_up_rate=config.investment_step_up_rate,
            rate_of_return=config.rate_of_return,
            inflation_rate=config.expense_inflation_rate
        )

        # Ensure the output directory exists
        output_path = os.path.join(output_dir, "financial_plan.xlsx")
        
        print(f"\nCreating Excel report at: {output_path}")
        if create_excel_report(retirement_df, investment_df, goal_dfs, summary_df, output_path):
            print("Excel report created successfully!")
            result = {
                "success": True,
                "required_monthly_investment": required_monthly_investment,
                "goal_monthly_investments": goal_monthly_investments,
                "final_retirement_corpus": final_retirement_corpus,
                "output_path": output_path
            }
            return json.loads(json.dumps(result, default=convert_to_serializable))
        else:
            return {"success": False, "message": "Failed to create the Excel report."}

    except Exception as e:
        print(f"Error in main function: {str(e)}")
        import traceback
        traceback.print_exc()
        return {"success": False, "error": str(e)}

if __name__ == "__main__":
    try:
        result = main()
        if result["success"]:
            print("\nFinancial plan generated successfully!")
            print(f"Required monthly investment: ₹{result['required_monthly_investment']:,.2f}")
            print(f"Final retirement corpus: ₹{result['final_retirement_corpus']:,.2f}")
            print(f"Excel report saved at: {result['output_path']}")
        else:
            print(f"\nError: {result.get('message', result.get('error', 'Unknown error'))}")
        sys.exit(0 if result["success"] else 1)
    except KeyboardInterrupt:
        print("\nProgram terminated by user.")
        sys.exit(1)
    except Exception as e:
        print(f"\nUnexpected error: {str(e)}")
        sys.exit(1)